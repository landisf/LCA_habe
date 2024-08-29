# Combine HABE data in ../habe20152017_hh_prepared_imputed.csv with LCA data in
# ../data-gwp/nfp73-ccl-preliminary-results-ipcc-gwp-april-2023.csv

import pandas as pd
import os
import pickle
import warnings
from openpyxl import load_workbook

# define paths
repo_root = os.path.abspath(os.getcwd())
filedir = os.path.join(repo_root,
                       'python')
out_dir = os.path.join(repo_root, 'intermediate_files')
config_dir = os.path.join(repo_root, 'config-aggregation')
in_dir = os.path.join(repo_root, 'intermediate_files')
# define filenames (including paths)
imputed_habe_file = os.path.join(in_dir,
                                 'habe20152017_imputed_withMZkm.csv')
categories_file = os.path.join(repo_root, 'data-elcom_froemelt',
                               'es8b01452_si_002.xlsx')
lca_file = os.path.join(repo_root, 'data-gwp',
                        'nfp73-ccl-preliminary-results-ipcc-gwp-april-2023.csv')
agg_conf_file = os.path.join(config_dir, 'aggregations.xlsx')


habe_out_file = os.path.join(out_dir, 'habe_lca.csv')
dict_out_file = os.path.join(out_dir, 'gwp_dict.pkl')

habe = pd.read_csv(imputed_habe_file,
                   sep=',',
                   header=0,
                   index_col=0)
lca = pd.read_csv(lca_file,
                  sep=',',
                  header=0,
                  index_col=0)

# categories_file describes the process of LCA analysis based on ecoinvent and others.
# I don't understand all of it but I use the contents for getting at the correspondance
# between English variable names and HABE categories.
a_hot_mess = pd.read_excel(categories_file,
                           sheet_name='Overview & LCA-Modeling',
                           header=2,
                           index_col=1)

# Translated name is the English variable name,
# Quantity code is the HABE variable code for quantities (codes starting with 'm')
# Variable code is the HABE variable code for expenditures (codes starting with 'a')
categories = pd.concat([a_hot_mess['Translated name'],
                        a_hot_mess['Quantity code'],
                        a_hot_mess['ConversionDem2FU']],
                       axis=1)
categories.index = a_hot_mess['Variable code']
categories.loc[categories['Quantity code'].notna()]
# Drop all entries that do not interest us:
# ConversionDem2FU is a proxy for 'Froemelt et al. used this in their LCA'
# (I checked visually that all lines with entries in ConversionDem2FU have at least
#  one non-zero entry in a column 'On X'; that is abstracting from the fact that
#  columns names are shifted/confused in the published file es8b01452_si_002.xlsx...)
categories = categories.loc[categories['ConversionDem2FU'].notna()]

# Prune LCA for entries that describe consumer goods, not households:
# lca['houshold_size'].isna() is a proxy for 'not a representative household'
lca = lca.loc[lca['houshold_size'].isna()].drop(columns=['houshold_size', 'gross_income',
                                                         'method', 'unit'])

# ----------------------------------------------------------------------------------------
# Fix entries of categories and lca for a one-to-one match between consumption goods
# ----------------------------------------------------------------------------------------

# There are trailing white spaces hanging out in categories. Let's strip them
categories['Translated name'] = categories['Translated name'].str.strip()

# categories['Translated name'] 'Fresh eggs' and 'Processed eggs' while
# lca.index contains 'Egg, national average, at farm gate' BESIDES
# 'Fresh eggs' and 'Processed eggs' but the latter two contain zeros

# Froemelt's categories_file suggests that national average at farm gate should be
# combined with medium voltage electricity and steam production from ecoinvent, but
# the latter two are to be multiplied with 0 -> I dare to make the simplification of
# setting fresh and processed eggs to "at farm gate"

lca.loc[lca.index == 'Egg, national average, at farm gate', :]
lca.loc[lca.index == 'Fresh eggs', :] = lca.loc[
    lca.index == 'Egg, national average, at farm gate', :].values
lca.loc[lca.index == 'Processed eggs', :] = lca.loc[
    lca.index == 'Egg, national average, at farm gate', :].values
lca.loc[lca.index == 'Fresh eggs', :]
lca.loc[lca.index == 'Processed eggs', :]
lca = lca.drop(labels='Egg, national average, at farm gate')
lca.loc[lca.index == 'Egg, national average, at farm gate', :]

# Fish: Froemelt et al seem to make this a combination of
# 0.6*bass+0.2*small trout+0.2*big trout
# (I'm Taking the 'Amount Act X' into account, but not 'CFL Act X'
#  or 'Conversion Dem2FU', since that should have been done in the LCA already.)
lca.loc[lca.index == 'Large trout, 2-4kg, conventional, at farm gate', 'mean']
lca.loc[lca.index == 'Small trout, 250-350g, conventional, at farm gate', 'mean']
lca.loc[
    lca.index == 'Sea bass or sea bream, 200-500g, conventional, in cage, at farm gate',
    'mean']
lca.loc[lca.index == 'Fish', 'mean']
lca.loc[lca.index == 'Fish', 'mean'] = (
    lca.loc[
        lca.index == 'Large trout, 2-4kg, conventional, at farm gate', 'mean'].values
    * 0.2
    + lca.loc[
        lca.index == 'Small trout, 250-350g, conventional, at farm gate', 'mean'].values
    * 0.2
    + lca.loc[
        lca.index == ('Sea bass or sea bream, 200-500g, conventional, in cage, '
                      + 'at farm gate'),
        'mean'].values
    * 0.6
)
lca.loc[lca.index == 'Fish', 'median'] = float("nan")
lca.loc[lca.index == 'Fish', 'std'] = float("nan")
lca.loc[lca.index == 'Fish', :]

lca = lca.drop(labels='Large trout, 2-4kg, conventional, at farm gate')
lca = lca.drop(labels='Small trout, 250-350g, conventional, at farm gate')
lca = lca.drop(
    labels='Sea bass or sea bream, 200-500g, conventional, in cage, at farm gate')

# LCA uses the term "education services", but I think it should be "education"
lca.loc[lca.index == "Education services", :]
lca = lca.rename(index={"Education services": "Education"})
lca.loc[lca.index == "Education services", :]  # empty
lca.loc[lca.index == "Education", :]           # new

# ----------------------------------------------------------------------------------------
# Check that all categories used by Froemelt et al. are avaible in our LCA and vice-versa
# ----------------------------------------------------------------------------------------
# Create numpy array (lose index 'variable names') of English product labels from LCA
products_lca = lca.index.to_numpy()
# Create numpy array of English product labels from categories
products_cat = categories['Translated name'].to_numpy()
# Check if there are parts of the LCA that cannot be used
for prod in products_lca:
    if prod not in products_cat:
        print('There is a problem. "', prod, '" is not in products_cat!', sep='')
# Check if there are parts of HABE that are not covered by LCA
for prod in products_cat:
    if prod not in products_lca:
        print('There is another problem: "', prod, '" is not in products_lca!', sep='')

# ----------------------------------------------------------------------------------------
# Check that all categories in our LCA can be allocated a HABE entry by Froemelt et al.
# ----------------------------------------------------------------------------------------

quant_lca = lca.loc[(lca["product_unit"] != "CHF") & (lca["product_unit"] != "unit")]
CHF_lca = lca.loc[lca["product_unit"] == "CHF"]

unit_lca = lca.loc[lca["product_unit"] == "unit"]

# These don't change anything; they just display some arrays and entries
quant_lca.index.to_numpy()
quant_lca
categories["Quantity code"].loc[categories["Translated name"] == 'Yoghurt'].values

# ----------------------------------------------------------------------------------------
# Check to what extent we have the quantities needed to apply LCA
# ----------------------------------------------------------------------------------------
# Go through those LCA entries that require physical quantities
# and check if categories has 'nan'.
# N.B.: LCA entries that neet quantities expressed in units (unit_lca)
#       are easily checked by hand and reveal no gaps.
for prod, row in quant_lca.iterrows():
    if prod not in products_cat:
        print('Cannot find "', prod, '"', sep='')
    else:
        quantid = str(categories["Quantity code"].loc[
            categories["Translated name"] == prod
        ][0])
        if quantid == "nan":
            print('"', prod, '" has quant variable "', quantid, '"', sep='')

# ----------------------------------------------------------------------------------------
# Let's go through the items that are flagged in above for loop:
#
# Things that can be handled:
# -----------------------------
# "Bakery products (sweet and salty)" has quant variable "nan"
#    => OK in HABE: M511104 in kg
habe['m511104']
# "Culinary herbs" has quant variable "nan"
#    => OK in HABE: M511703 in kg
habe['m511703']
# "Fresh eggs" has quant variable "nan"
#    => OK in HABE:   M511409 (frisch) in kg
habe['m511409']
# "Processed eggs" has quant variable "nan"
#    => OK in HABE:   M511410 (verarbeitet) in kg
habe['m511410']
# "Other sugary or cocoa-based foods" has quant variable "nan"
#    => OK: M511807 in kg
habe['m511807']
# "Ready-made foods" has quant variable "nan"
#    => OK: M511905 in kg
habe['m511905']
# "Sandwich" has quant variable "nan"
#    => OK: M511105 in kg
habe['m511105']
# "Sauces, seasonings and condiments" has quant variable "nan"
#    => OK: M511901 in kg
habe['m511901']
# "Soups and bouillons" has quant variable "nan"
#    => OK: M511903 in kg
habe['m511903']
# "Sweets and chewing gum" has quant variable "nan"
#    => OK: M511805 in kg
habe['m511805']
# "Vegetarian soy products" has quant variable "nan"
#    => OK: M511906 in kg
habe['m511906']
#    "Bus, tickets and travelcards" (person kilometers) has quant variable "nan"
habe['pkm_bus']
#    "Train, tickets and travelcards" (person kilometers) has quant variable "nan"
habe['pkm_train']
#    "Tram, tickets and travelcards" (person kilometers) has quant variable "nan"
habe['pkm_tram']
#    "Bicycles" (person kilometers) has quant variable "nan"
habe['pkm_bike']

# Genuine NaNs:
# ---------------
# Froemelt used Microcensus to get at person kilometers
# Maybe also Microcensus?:

# I guess the following are borderline irrelevant, but could consult with PSI/Froemelt's
# methodology to see if these could also be derived from CHF
# (Froemelt et al. paper specifically mentions the categories as being among those where
# CHF->kg conversion had to made; how did they do it? They cite another paper...)
#    "Newspapers and periodicals" (kg) has quant variable "nan"
#    "Soaps and foam baths" (kg) has quant variable "nan"

# ----------------------------------------------------------------------------------------
# Apply what we just learned: append column to LCA
# ----------------------------------------------------------------------------------------
lca.insert(len(lca.columns), "habe_item", float("nan"))
lca2 = lca.copy()
lca = lca2.copy()
for prod in lca.index:
    # print(type(prod), prod, lca['product_unit'][prod])
    if prod not in products_cat:
        print('Cannot find "', prod, '"', sep='')
    else:
        # Manually add the stuff that's not given by categories
        match prod:
            case "Bus, tickets and travelcards":
                lca.loc[prod, 'habe_item'] = 'pkm_bus'
            case "Train, tickets and travelcards":
                lca.loc[prod, 'habe_item'] = 'pkm_train'
            case "Tram, tickets and travelcards":
                lca.loc[prod, 'habe_item'] = 'pkm_tram'
            case "Bicycles":
                lca.loc[prod, 'habe_item'] = 'pkm_bike'
            case "Bakery products (sweet and salty)":
                lca.loc[prod, 'habe_item'] = 'm511104'
            case "Culinary herbs":
                # lca['habe_item'][prod] = 'm511703'
                lca.loc[prod, 'habe_item'] = 'm511703'
            case "Fresh eggs":
                lca.loc[prod, 'habe_item'] = 'm511409'
            case "Processed eggs":
                lca.loc[prod, 'habe_item'] = 'm511410'
            case "Other sugary or cocoa-based foods":
                lca.loc[prod, 'habe_item'] = 'm511807'
            case "Ready-made foods":
                lca.loc[prod, 'habe_item'] = 'm511905'
            case "Sandwich":
                lca.loc[prod, 'habe_item'] = 'm511105'
            case "Sauces, seasonings and condiments":
                lca.loc[prod, 'habe_item'] = 'm511901'
            case "Soups and bouillons":
                lca.loc[prod, 'habe_item'] = 'm511903'
            case "Sweets and chewing gum":
                lca.loc[prod, 'habe_item'] = 'm511805'
            case "Vegetarian soy products":
                lca.loc[prod, 'habe_item'] = 'm511906'
            case _:
                if lca['product_unit'][prod] == "CHF":
                    # First monetary stuff:
                    exp_id = (categories.loc[
                        categories["Translated name"] == prod].index[0])
                    lca.loc[prod, 'habe_item'] = exp_id
                elif lca['product_unit'][prod] == "unit":
                    # Then things measured in units owned:
                    unit_id = (categories.loc[
                        categories["Translated name"] == prod].index[0])
                    lca.loc[prod, 'habe_item'] = unit_id
                else:
                    # Now the things that require physical quants
                    quant_id = (categories["Quantity code"].loc[
                        categories["Translated name"] == prod][0])
                    lca.loc[prod, 'habe_item'] = quant_id

# See, this was a success for self created habe entries with person kilometers (pkm):
lca['habe_item'].loc['Bicycles']
habe['pkm_bike']
lca['habe_item'].loc['Bus, tickets and travelcards']
habe['pkm_bus']
lca['habe_item'].loc['Tram, tickets and travelcards']
habe['pkm_tram']
lca['habe_item'].loc['Train, tickets and travelcards']
habe['pkm_train']
# See, this was a success for unit_lca:
lca['habe_item'].loc['Printers (incl. multifunctional printers)']
habe['cg_noprinters']
lca['habe_item'].loc['Desktop computers']
habe['cg_nodesktoppcs']
print(lca['habe_item'].loc['Portable computers'])
habe['cg_nolaptops']
# And for quant_lca?
# lca['habe_item'].loc['Newspapers and periodicals']
# for prod, row in quant_lca.iterrows():
#     print(lca['habe_item'].loc[prod])


# Check that the NaNs written above are for known offenders only
lca.loc[lca['habe_item'] != lca['habe_item'], 'habe_item']

# At this point, lca knows which HABE variables have to be multiplied with GWPs
# Things for which we have a GPW-LCA, that are not yet matched with HABE spending:
# Fishes (in kg), IT devices (their unit is 'unit'),
# Desktop computers	cg_nodesktoppcs
# Portable computers	cg_nolaptops
# Printers (incl. multifunctional printers)	cg_noprinters

# - Transportation? LCA for bicycles, bus, train, tram in person kilometer
#   1. How to infer person kilometers from öV spending? bikes???
#   2. What about car stocks? Froemelt et al. are awkwardly silent about this.
#      (They use stocks for IT equipment, spending for white goods, but neither for cars.)
#      Car wear/annualized car GWPs are likely incorporated in car kilometers linked to
#      fuel demand!
# - Nebenwohnsitze (secondary residence)? Froemelt et al. add this to primary numbers
# - White goods are under running consumption, but IT as durable goods???
#   Froemelt et al.'s paper is silent on this distinction. This might be driven by
#   units available in ecoinvent, but it would be good to at least know if things are
#   annualized over lifetime for IT devices.

# Other gaps in Froemelt: insurance, health insurance,
# (yet they include 'other insurance' and 'life insurance'...)

# ----------------------------------------------------------------------------------------
# Let's multiply some things and write to HABE
# ----------------------------------------------------------------------------------------
# This comes close to what I want to do, but I need to figure out what's up with NaNs...
gwp_dict = {}                   # type(gwp_dict) returns dict
gwp_list = []                 # type(gwp_list) returns list
with warnings.catch_warnings():
    # warnings.filterwarnings can filter by:
    # - message='string' where string is a regex or substring of the messages to filter
    # - lineno
    # - module
    # - category
    warnings.filterwarnings("ignore", message='DataFrame is highly fragmented.')
    for prod in lca.index:
        lca_label = lca['habe_item'][prod]
        # hack for checking nan-ness
        if lca_label != lca_label:
            print('Caught a NaN:', prod)
        # elif lca_label == 'nan':
        #     # This happens for bicycles, newspapers, soaps and tickets:
        #     # Items that can be found in categories, but don't have quantity entries
        #     print('"nan" found for prod "', prod, '".', sep='')
        else:
            # gwp_label = 'gwp'+lca_label.lstrip('amxcg_no')
            gwp_label = 'gwp'+lca_label.lstrip('amxcgpkm')
            gwp_list.append(gwp_label)
            gwp_dict[gwp_label] = prod
            if gwp_label != 'gwpnan':
                habe.insert(len(habe.columns), gwp_label, float("nan"))
                # Here I multiply HABE's consumption units with LCA's GWP per unit
                habe[gwp_label] = habe[lca_label]*lca['mean'][prod]

gwp_list.sort()
gwp_list
gwp_df = pd.DataFrame(gwp_list, columns=['labels'])


book = load_workbook(agg_conf_file)
book.sheetnames
book.worksheets

writer = pd.ExcelWriter(agg_conf_file,
                        engine='openpyxl',
                        mode='a',
                        if_sheet_exists='replace')

gwp_df.transpose().to_excel(writer,
                            sheet_name='gwp_list',
                            header=False,
                            index=False)
writer.close()

# Look up some things:
gwp_dict['gwp_nolaptops']
habe['gwp_nolaptops']
# Python suggests this for defragmentation of data frame:
habe = habe.copy()

# HABE with GWP per household and consumption category (units consumed * GWP per unit)
habe.to_csv(habe_out_file)

with open(dict_out_file, 'wb') as f:
    pickle.dump(gwp_dict, f)

exit()

# ----------------------------------------------------------------------------------------
# Going through Froemelt's SI: Covered by supercategories in Froemelt
# ----------------------------------------------------------------------------------------
# Fish (a5113) covers all its subcategories .01--.05
# Wine (a5212) covers all its subcategories .01--.09
# Tobacco (a5220) covers all its subcategories .01--.02
# Restaurants, cafés and bars (a5311) -.03
# Self-service restaurants and take-aways	a5312 covers -.03
# Canteens	a5313 covers -.03
# Accommodation services	a532 (same name as a5320) covers  5320.01-.02
# Garments for men 	a5612 covers .01--.06
# Garments for women	a5613 covers .01--.07
# Garments for children (between 0 and 13 years)	a5614 covers --.06
# (Footwear for men	a5621 covers a562100 BUT IS CALLED THE SAME!)
# (Footwear for women	a5622 covers a562200 BUT IS CALLED THE SAME!)
# (Footwear for children and babies	a5623 covers a562300 BUT IS CALLED THE SAME!)
# Repair and hire of footwear	a5624 covers a562400
# Maintenance and repair of the dwelling a573 (same name as a5730) covers a5730.01-.02
# Furniture and furnishings, carpets and other floor coverings incl. repairs	a581
#    (a5810) covers a5810.01-.05
# Household textiles	a582 (a5820) a5820.01-.03
# Tools and equipment without engines for house and garden	a5841 covers
#    .01-.02
# (Machines with engines for house and garden	a5842 has same name as a584200)
# Medical services and hospital services	a612 (a6120) covers .01-.05
# (Bicycles	a6213 has same name as a621300)
# Passenger transport by sea and inland waterway	a6224 covers a622400
# Other purchased transport services 	a6226 covers a622600
# Fixed line	a6322 covers .01-.04
# Mobile telephony	a6323 covers .01-.03
# Services of internet providers	a6324 covers .01-0.2
# Sound carriers, data carriers, videotapes and films	a6614 covers .01-.02
# (Durable goods for recreation and sports	a6621 has same name as a662100)
# Games, toys and hobbies	a6622 covers .01-.04
# Equipment for sports and camping	a6623 covers .01-.03
# (Plants and other non-durable goods for gardening	a6624 has same name as a662400)
# Stakes	a6633 covers a663300
# (Books and brochures	a6641 has same name as a664100)
# Newspapers and periodicals	a6642 covers .01-.02
# Education	a67 covers a670, a6700, and .01-.04
# Social protection services	a6831 covers .00
# (Financial services	a6832 has same name as a683200)
# Dues to organisations and associations	a6833 covers .01-.05
# (Other services and losses relating to renting	a6834 has same name as .00)
# Other insurance premiums	a42  covers a420, a4201 (.01-.04), a4202 (.01-.02),
#                                           a4203 (.01-.04)
# Fees	a43 covers a430, a4300 .01-.06
# Premiums for life insurance	a80 (same name as a81, a810, a8100) covers .01-.02

# ----------------------------------------------------------------------------------------
# Not covered in Froemelt
# ----------------------------------------------------------------------------------------
# Private invitation	a5314, including subcategories .01-.03
# Net rent and mortgage interest of principal residence	a5711, including a571100
# Utility lump sum of principal residence	a571201
# Rent, mortgage interest, utility costs and energy of secondary residences	a572
#   including a5721, a5722, a5723
#             a572100, a572200, a572300
# Cars	a6211 a6211.01-.02
# Motor cycles, scooters and mopeds	a6212 (a621200)
# Spare parts and accessories for vehicles	a6214 a6214.01--.02
# Lubricants and other care products 	a621503
# Service and repairs of vehicles	a6216 (a621600)
# Other private vehicle services	a6217  and a6217.01-.05
# Combined passenger transport	a6225 and .01-.04
# Computer, office appliances and other peripherals	a6613 and .01-.02
# Compulsory transfer expenditure	a30
# Social security	a31 (a310, a3100, .01-.05)
# Taxes	a32 (a320, a3200, .01-.04)
# Health insurance: Premiums of basic insurance	a33 (a330, a3300, .01-.02)
# Monetary transfer expenditures to other households	a35
# Monetary transfer expenditures to other households	a36 (a360, a3600, .01-.03)
# Other insurances, fees and transfers	a40
# Health insurance: Premiums for supplementary insurance	a41 (a410, a4100, .01-.02)
# Donations, gifts and invitations	a44 (lots...)
#


# ----------------------------------------------------------------------------------------
# INSIGHTS
#
# LCA has entries for 'Eggs, national average...' and zeros for 'fresh eggs' and
# 'processed eggs'. Here I just overwrote zeros with national average and dropped the
# latter.
#
# LCA has zeros for 'Fish' which is a four digit category
# Froemelt's categories_file suggest that 'Fish' should be constructed from 2 trouts and
# bass.
#
# For electronics, things look patchy even in Froemelt's categories_file
# PSI seems to resort to owned number of devices -> how do they do that? how to annualize?
# Available are: 'Desktop computers', 'Portable computers', and
# 'Printers (incl. multifunctional printers)'
#
# Cars: do not appear to be a thing in LCA_file. Froemelt et al. seem to be clueless, too.
#
# CONSEQUENCES
# - check, for which level categories (four digit/five digit etc.) things are available
# - learn how Froemelt or PSI constructed Fish and similar from trouts and bass
# - figure out, what to do with electronics.
# - Rename 'Education services' to 'Education'? Digits!
# - Get an overview over uncovered categories and write caveats section...
#   -> reverse for-loop to see what categories are uncovered in LCA.
# ----------------------------------------------------------------------------------------
# There is a problem. Desktop computers is not in products_cat!
# There is a problem.
#     Large trout, 2-4kg, conventional, at farm gate is not in products_cat!
# There is a problem. Portable computers is not in products_cat!
# There is a problem. Printers (incl. multifunctional printers) is not in products_cat!
# There is a problem.
#     Sea bass or sea bream, 200-500g, conventional, in cage, at farm gate
#     is not in products_cat!
# There is a problem.
#     Small trout, 250-350g, conventional, at farm gate is not in products_cat!
